import os

from PyQt5.QtWidgets import QDialog, QVBoxLayout, QPushButton, QMessageBox, QInputDialog
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class SettingsDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Настройки")
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Добавление кнопок "Добавить продукт на линию" и "Удалить продукт с линии"
        add_product_button = QPushButton("Добавить продукт на линию")
        add_product_button.clicked.connect(self.add_product_to_line)
        layout.addWidget(add_product_button)

        remove_product_button = QPushButton("Удалить продукт с линии")
        remove_product_button.clicked.connect(self.remove_product_from_line)
        layout.addWidget(remove_product_button)

        # Добавление кнопки "Настройка состава продукта"
        product_composition_button = QPushButton("Настройка состава продукта")
        product_composition_button.clicked.connect(self.open_product_composition_dialog)
        layout.addWidget(product_composition_button)

        # Добавление кнопки "Расчет остатков материалов"
        calculate_stock_button = QPushButton("Расчет остатков материалов")
        calculate_stock_button.clicked.connect(self.calculate_material_stock)
        layout.addWidget(calculate_stock_button)

        # Атрибут для хранения добавленных материалов
        self.added_materials = []

    def add_product_to_line(self):
        lines = ["Линия 1,5", "Линия 0,5", "Линия 1,5 сладкая", "Линия 0,5 сладкая", "Линия 5 литров", "Линия 19 литров"]
        selected_line, ok = QInputDialog.getItem(self, "Выберите линию", "Выберите линию:", lines, 0, False)
        if not ok:
            return

        types = ["Газированный", "Не газированный"]
        selected_type, ok = QInputDialog.getItem(self, "Выберите тип продукта", "Выберите тип продукта:", types, 0, False)
        if not ok:
            return

        wells = ["1РЭ", "Суздальская", "а21", "а22"]
        selected_well, ok = QInputDialog.getItem(self, "Выберите скважину", "Выберите скважину:", wells, 0, False)
        if not ok:
            return

        name, ok = QInputDialog.getText(self, "Введите наименование", "Введите наименование продукта:")
        if not ok or not name:
            return

        gtin, ok = QInputDialog.getText(self, "Введите GTIN", "Введите GTIN продукта:")
        if not ok or not gtin:
            return

        # Настройка состава продукта
        composition = self.setup_product_composition(name)
        if not composition:
            QMessageBox.warning(self, "Ошибка", "Не удалось настроить состав продукта.")
            return

        filename = "production.xlsx"
        if not os.path.exists(filename):
            wb = Workbook()
            wb.remove(wb.active)
            for line in lines:
                wb.create_sheet(line)
            wb.save(filename)

        wb = load_workbook(filename)
        if selected_line in wb.sheetnames:
            ws = wb[selected_line]
        else:
            QMessageBox.warning(self, "Ошибка", f"Линия {selected_line} не найдена.")
            return

        ws.append([name, gtin, selected_well, selected_type])
        wb.save(filename)
        wb.close()

        QMessageBox.information(self, "Успех", f"Продукт '{name}' успешно добавлен на линию {selected_line}.")

    def remove_product_from_line(self):
        lines = ["Линия 1,5", "Линия 0,5", "Линия 1,5 сладкая", "Линия 0,5 сладкая", "Линия 5 литров", "Линия 19 литров"]
        selected_line, ok = QInputDialog.getItem(self, "Выберите линию", "Выберите линию:", lines, 0, False)
        if not ok:
            return

        filename = "production.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с продукцией не найден.")
            return

        wb = load_workbook(filename)
        if selected_line not in wb.sheetnames:
            QMessageBox.warning(self, "Ошибка", f"Линия {selected_line} не найдена.")
            return

        ws = wb[selected_line]

        products = [(cell.value, idx) for idx, cell in enumerate(ws['A'], start=1)]
        product_names = [name for name, _ in products]

        product_name, ok = QInputDialog.getItem(self, "Выберите продукт", "Выберите продукт для удаления:", product_names, 0, False)
        if not ok:
            return

        idx = next((idx for name, idx in products if name == product_name), None)
        if idx is None:
            QMessageBox.warning(self, "Ошибка", "Продукт не найден.")
            return

        ws.delete_rows(idx)
        wb.save(filename)
        wb.close()

        QMessageBox.information(self, "Успех", f"Продукт '{product_name}' успешно удален с линии {selected_line}.")

    def open_product_composition_dialog(self):
        QMessageBox.information(self, "Предупреждение", "Функционал настройки состава продукта в разработке.")

    def setup_product_composition(self, product_name):
        composition = []

        # Создаем или открываем файл product_composition.xlsx
        filename = "product_composition.xlsx"
        if not os.path.exists(filename):
            wb = Workbook()
        else:
            wb = load_workbook(filename)

        # Создаем лист для текущего продукта или открываем существующий
        if product_name in wb.sheetnames:
            ws = wb[product_name]
        else:
            ws = wb.create_sheet(product_name)

        while True:
            available_materials = self.get_available_materials(product_name)  # Получаем доступные материалы
            if not available_materials:  # Если список пуст, выходим из цикла
                break

            material, ok = QInputDialog.getItem(self, "Выберите материал",
                                                "Выберите материал, содержащийся в продукте:", available_materials, 0,
                                                False)
            if not ok or not material:
                break

            quantity, ok = QInputDialog.getDouble(self, "Введите количество",
                                                  f"Введите количество материала '{material}':", min=0.1)
            if not ok:
                break

            composition.append((material, quantity))
            self.added_materials.append(
                (product_name, material))  # Добавляем выбранный материал в список уже добавленных

            # Записываем материал и его количество на лист текущего продукта
            ws.append([material, quantity])

        # Сохраняем и закрываем файл
        wb.save(filename)
        wb.close()

        return composition

    def get_available_materials(self, product_name):
        # Получение списка доступных для выбора материалов
        filename = "materials.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с материалами не найден.")
            return []

        wb = load_workbook(filename)
        ws = wb.active

        materials = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            material = row[0]
            if material not in self.get_added_materials(product_name):  # Исключаем уже добавленные материалы
                materials.append(material)
        wb.close()
        return materials

    def get_added_materials(self, product_name):
        # Получение списка уже добавленных материалов для конкретного продукта
        added_materials = set()
        for name, material in self.added_materials:
            if name == product_name:
                added_materials.add(material)
        return added_materials

    def calculate_material_stock(self):
        # Загрузить данные о материалах, выпущенной продукции и составе продуктов
        materials_filename = "materials.xlsx"
        parties_filename = "parties.xlsx"
        composition_filename = "product_composition.xlsx"

        # Проверяем наличие файлов
        if not os.path.exists(materials_filename) or not os.path.exists(parties_filename) or not os.path.exists(
                composition_filename):
            QMessageBox.warning(self, "Ошибка", "Файлы с данными не найдены.")
            return

        # Загружаем книги Excel
        materials_wb = load_workbook(materials_filename)
        parties_wb = load_workbook(parties_filename)
        composition_wb = load_workbook(composition_filename)

        materials_ws = materials_wb.active
        parties_ws = parties_wb.active

        # Проходим по каждой партии в файле parties.xlsx
        for row in parties_ws.iter_rows(min_row=2, max_col=5, values_only=True):
            _, product_name, _, _, quantity = row
            if product_name is None or quantity is None:
                continue

            # Получаем лист с составом продукта
            product_sheet = composition_wb[product_name]

            # Обрабатываем состав продукта, представленный в строках
            for comp_row in product_sheet.iter_rows(min_row=1, max_col=2, values_only=True):
                material, comp_quantity = comp_row
                # Получаем текущий остаток материала
                current_stock = 0
                for mat_row in materials_ws.iter_rows(min_row=2, max_col=3, values_only=True):
                    mat_name, _, stock = mat_row
                    if mat_name == material:
                        current_stock = stock
                        break

                # Обновляем остаток материала
                new_stock = current_stock - (comp_quantity * quantity)
                for mat_row in materials_ws.iter_rows(min_row=2, max_col=3):
                    mat_name, _, stock = mat_row
                    if mat_name.value == material:
                        stock.value = new_stock
                        break

        # Сохраняем обновленные остатки материалов
        materials_wb.save(materials_filename)
        materials_wb.close()
        parties_wb.close()
        composition_wb.close()

        # Оповещение об успешном обновлении остатков материалов
        QMessageBox.information(self, "Успех", "Остатки материалов на складе успешно обновлены.")

if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication
    app = QApplication(sys.argv)
    dialog = SettingsDialog()
    dialog.show()
    sys.exit(app.exec_())
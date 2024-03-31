import os
import sys
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QMessageBox
from openpyxl import load_workbook, Workbook
import win32api
import win32print
from datetime import datetime

class CheckDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Сверка")

        layout = QVBoxLayout()
        self.setLayout(layout)

        # Добавление кнопки "Показать все материалы"
        show_all_materials_button = QPushButton("Показать все материалы")
        show_all_materials_button.clicked.connect(self.show_all_materials)
        layout.addWidget(show_all_materials_button)

        # Добавление кнопки "Сохранить и распечатать"
        save_and_print_button = QPushButton("Сохранить и распечатать")
        save_and_print_button.clicked.connect(self.save_and_print)
        layout.addWidget(save_and_print_button)

        # Таблица для отображения материалов
        self.table = QTableWidget()
        layout.addWidget(self.table)

    def show_all_materials(self):
        filename = "materials.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с материалами не найден.")
            return

        wb = load_workbook(filename)
        ws = wb.active

        rows = ws.max_row
        cols = ws.max_column

        self.table.setRowCount(rows)
        self.table.setColumnCount(cols)
        headers = ["Наименование", "Единица измерения", "Количество"]
        self.table.setHorizontalHeaderLabels(headers)

        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                item = QTableWidgetItem(str(ws.cell(row=row, column=col).value))
                self.table.setItem(row - 1, col - 1, item)

        wb.close()

    def save_and_print(self):
        # Создание имени файла с текущей датой
        current_date = datetime.now().strftime("%Y-%m-%d")
        filename = f"materials_export_{current_date}.xlsx"
        desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop', filename)

        # Сохранение данных из таблицы в файл
        wb = Workbook()
        ws = wb.active

        rows = self.table.rowCount()
        cols = self.table.columnCount()

        for row in range(rows):
            for col in range(cols):
                item = self.table.item(row, col)
                if item is not None:
                    ws.cell(row=row + 1, column=col + 1, value=item.text())

        wb.save(desktop_path)

        # Печать файла
        win32api.ShellExecute(0, "print", desktop_path, f'"{win32print.GetDefaultPrinter()}"', ".", 0)

        QMessageBox.information(self, "Успех", f"Данные успешно сохранены и отправлены на печать.")

if __name__ == "__main__":
    from PyQt5.QtWidgets import QApplication

    app = QApplication(sys.argv)
    dialog = CheckDialog()
    dialog.show()
    sys.exit(app.exec_())
import os

from PyQt5.QtWidgets import QDialog, QVBoxLayout, QPushButton, QMessageBox, QInputDialog, QTableWidget, QTableWidgetItem
from openpyxl import Workbook, load_workbook


class MaterialWarehouseDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Склад материалов")
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Добавление кнопок "Добавить материал на склад", "Удалить материал со склада" и "Корректировать остатки"
        add_material_button = QPushButton("Добавить материал на склад")
        add_material_button.clicked.connect(self.add_material_to_warehouse)
        layout.addWidget(add_material_button)

        remove_material_button = QPushButton("Удалить материал со склада")
        remove_material_button.clicked.connect(self.remove_material_from_warehouse)
        layout.addWidget(remove_material_button)

        adjust_stock_button = QPushButton("Корректировать остатки")
        adjust_stock_button.clicked.connect(self.adjust_stock)
        layout.addWidget(adjust_stock_button)

        show_all_materials_button = QPushButton("Показать все материалы")
        show_all_materials_button.clicked.connect(self.show_all_materials)
        layout.addWidget(show_all_materials_button)

    def add_material_to_warehouse(self):
        name, ok = QInputDialog.getText(self, "Введите наименование", "Введите наименование материала:")
        if not ok or not name:
            return

        units = ["шт", "л", "г", "кг"]  # Список доступных единиц измерения
        unit, ok = QInputDialog.getItem(self, "Выберите единицу измерения", "Выберите единицу измерения материала:",
                                        units, 0, False)
        if not ok or not unit:
            return

        quantity, ok = QInputDialog.getInt(self, "Введите количество", "Введите количество материала:")
        if not ok or quantity <= 0:
            QMessageBox.warning(self, "Ошибка", "Введите корректное количество материала.")
            return

        filename = "materials.xlsx"
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["Наименование", "Единица измерения", "Количество"])  # Добавляем заголовок
            wb.save(filename)
            wb.close()

        wb = load_workbook(filename)
        ws = wb.active

        # Проверяем, есть ли уже заголовок
        if ws['A1'].value != "Наименование" or ws['B1'].value != "Единица измерения" or ws['C1'].value != "Количество":
            ws.insert(1, ["Наименование", "Единица измерения", "Количество"])

        ws.append([name, unit, quantity])
        wb.save(filename)
        wb.close()

        QMessageBox.information(self, "Успех", f"Материал '{name}' успешно добавлен на склад.")

    def remove_material_from_warehouse(self):
        filename = "materials.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с материалами не найден.")
            return

        wb = load_workbook(filename)
        ws = wb.active

        materials = [(cell.value, idx) for idx, cell in enumerate(ws['A'], start=1)]
        material_names = [name for name, _ in materials]

        material_name, ok = QInputDialog.getItem(self, "Выберите материал", "Выберите материал для удаления:", material_names, 0, False)
        if not ok:
            return

        idx = next((idx for name, idx in materials if name == material_name), None)
        if idx is None:
            QMessageBox.warning(self, "Ошибка", "Материал не найден.")
            return

        ws.delete_rows(idx)
        wb.save(filename)
        wb.close()

        QMessageBox.information(self, "Успех", f"Материал '{material_name}' успешно удален со склада.")

    def adjust_stock(self):
        filename = "materials.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с материалами не найден.")
            return

        wb = load_workbook(filename)
        ws = wb.active

        materials = [(cell.value, idx) for idx, cell in enumerate(ws['A'], start=1)]
        material_names = [name for name, _ in materials]

        material_name, ok = QInputDialog.getItem(self, "Выберите материал", "Выберите материал для корректировки остатков:", material_names, 0, False)
        if not ok:
            return

        idx = next((idx for name, idx in materials if name == material_name), None)
        if idx is None:
            QMessageBox.warning(self, "Ошибка", "Материал не найден.")
            return

        new_quantity, ok = QInputDialog.getInt(self, "Введите новое количество", f"Текущее количество: {ws.cell(row=idx, column=3).value}\nВведите новое количество материала '{material_name}':")
        if not ok or new_quantity <= 0:
            QMessageBox.warning(self, "Ошибка", "Введите корректное количество материала.")
            return

        ws.cell(row=idx, column=3).value = new_quantity
        wb.save(filename)
        wb.close()

        QMessageBox.information(self, "Успех", f"Остатки материала '{material_name}' успешно скорректированы.")

    def show_all_materials(self):
        filename = "materials.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с материалами не найден.")
            return

        wb = load_workbook(filename)
        ws = wb.active

        rows = ws.max_row
        cols = ws.max_column

        # Создание таблицы и добавление данных в нее
        table = QTableWidget()
        table.setRowCount(rows)
        table.setColumnCount(cols)
        headers = ["Наименование", "Единица измерения", "Количество"]
        table.setHorizontalHeaderLabels(headers)

        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                item = QTableWidgetItem(str(ws.cell(row=row, column=col).value))
                table.setItem(row - 1, col - 1, item)

        # Отображение таблицы в диалоговом окне
        layout = QVBoxLayout()
        layout.addWidget(table)

        dialog = QDialog()
        dialog.setWindowTitle("Остатки материалов")
        dialog.setLayout(layout)
        dialog.exec_()

        wb.close()

if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication
    app = QApplication(sys.argv)
    dialog = MaterialWarehouseDialog()
    dialog.show()
    sys.exit(app.exec_())
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QPushButton, QLabel, QComboBox, QMessageBox
from PyQt5.QtCore import QDate, Qt
from datetime import datetime

from openpyxl.workbook import Workbook

class ProductionPartyDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Создать партию")
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Добавление метки и выпадающего списка для выбора продукта
        product_label = QLabel("Выберите продукт:")
        layout.addWidget(product_label)
        self.product_combo_box = QComboBox()
        layout.addWidget(self.product_combo_box)

        # Добавление метки и выпадающего списка для выбора линии
        line_label = QLabel("Выберите линию:")
        layout.addWidget(line_label)
        self.line_combo_box = QComboBox()
        layout.addWidget(self.line_combo_box)

        # Добавление кнопки "Создать партию"
        create_party_button = QPushButton("Создать партию")
        create_party_button.clicked.connect(self.create_party)
        layout.addWidget(create_party_button)

        # Заполнение выпадающего списка линий
        self.populate_line_combo_box()

        # Обновление списка продуктов при выборе линии
        self.line_combo_box.currentIndexChanged.connect(self.populate_product_combo_box)

    def populate_line_combo_box(self):
        filename = "production.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с продукцией не найден.")
            return

        wb = load_workbook(filename)
        lines = set()
        for sheet in wb.sheetnames:
            lines.add(sheet)
        wb.close()

        self.line_combo_box.clear()
        self.line_combo_box.addItems(sorted(lines))
        self.populate_product_combo_box()

    def populate_product_combo_box(self):
        selected_line = self.line_combo_box.currentText()
        filename = "production.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с продукцией не найден.")
            return

        wb = load_workbook(filename)
        ws = wb[selected_line]
        products = [cell.value for cell in ws['A'] if cell.value]
        self.product_combo_box.clear()
        self.product_combo_box.addItems(sorted(products))

    def create_month_sheet_if_not_exists(self, filename, month_year):
        wb = load_workbook(filename)
        if month_year not in wb.sheetnames:
            ws = wb.create_sheet(month_year)
            ws.append(["Название линии", "Название продукта", "Партия", "Дата", "Количество"])
            wb.save(filename)
        wb.close()

    def create_party(self):
        selected_product = self.product_combo_box.currentText()
        selected_line = self.line_combo_box.currentText()
        selected_date = QDate.currentDate().toString("dd.MM.yyyy")

        filename = "parties.xlsx"
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["Название линии", "Название продукта", "Партия", "Дата", "Количество"])
            wb.save(filename)
            wb.close()

        wb = load_workbook(filename)

        # Получение текущего месяца и года
        current_date = datetime.now()
        current_month_year = current_date.strftime("%B %Y")

        # Проверка и создание листа для текущего месяца и года
        self.create_month_sheet_if_not_exists(filename, current_month_year)

        ws = wb[current_month_year]

        max_party_number = 0
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            if row[2] and isinstance(row[2], int) and row[2] > max_party_number:
                max_party_number = row[2]

        next_party_number = max_party_number + 1

        ws.append([selected_line, selected_product, next_party_number, selected_date, 0])
        wb.save(filename)
        wb.close()

        QMessageBox.information(self, "Успех", f"Создана партия {next_party_number} для продукта '{selected_product}' на линии '{selected_line}' с датой производства {selected_date}.")
if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication
    app = QApplication(sys.argv)
    dialog = ProductionPartyDialog()
    dialog.exec_()
    sys.exit(app.exec_())
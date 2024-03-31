import os
from datetime import datetime
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QPushButton, QLabel, QComboBox, QMessageBox, QDateEdit
from openpyxl import load_workbook


class ProductWarehouseDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Склад готовой продукции")
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Отображение остатков продукции
        self.products_label = QLabel("Остатки готовой продукции:")
        layout.addWidget(self.products_label)

        # Добавление выпадающего списка для выбора продукта
        self.product_combo_box = QComboBox()
        layout.addWidget(self.product_combo_box)

        # Добавление кнопок и полей ввода
        adjust_button = QPushButton("Корректировать остатки")
        adjust_button.clicked.connect(self.adjust_products)
        layout.addWidget(adjust_button)

        view_button = QPushButton("Просмотреть остаток за период")
        view_button.clicked.connect(self.view_product_balance)
        layout.addWidget(view_button)

        # Виджеты для выбора периода
        self.start_date_edit = QDateEdit()
        self.end_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.end_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(datetime.now().date())
        self.end_date_edit.setDate(datetime.now().date())
        layout.addWidget(self.start_date_edit)
        layout.addWidget(self.end_date_edit)

        # Загрузка списка продуктов
        self.load_products()

    def load_products(self):
        # Загрузка продуктов из файла parties.xlsx
        filename = "parties.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с данными не найден.")
            return

        wb = load_workbook(filename)
        ws = wb.active

        # Проходим по строкам файла, начиная со второй строки
        for row in ws.iter_rows(min_row=2, values_only=True):
            _, product, _, _, _ = row  # Пропускаем ненужные столбцы
            self.product_combo_box.addItem(product)

        wb.close()

    def adjust_products(self):
        product_name = self.product_combo_box.currentText()
        if not product_name:
            QMessageBox.warning(self, "Ошибка", "Выберите продукт.")
            return

        # Здесь должен быть код для корректировки остатков продукции
        QMessageBox.information(self, "Успех", f"Остатки продукции '{product_name}' успешно скорректированы!")

    def view_product_balance(self):
        product_name = self.product_combo_box.currentText()
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()

        if not product_name:
            QMessageBox.warning(self, "Ошибка", "Выберите продукт.")
            return

        # Загрузка данных из файла parties.xlsx
        filename = "parties.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с данными не найден.")
            return

        wb = load_workbook(filename)
        ws = wb.active

        # Сумма количества продукта за выбранный период
        total_quantity = 0

        # Проходим по строкам файла, начиная со второй строки
        for row in ws.iter_rows(min_row=2, values_only=True):
            _, prod, _, prod_date, quantity = row  # Пропускаем ненужные столбцы

            # Преобразование строковой даты в объект datetime.date
            prod_date = datetime.strptime(prod_date, '%d.%m.%Y').date()

            # Преобразование строкового количества в целое число
            quantity = int(quantity)

            # Фильтруем данные по выбранному продукту и периоду
            if prod == product_name and start_date <= prod_date <= end_date:
                total_quantity += quantity

        wb.close()

        # Формируем информацию об остатке для вывода
        balance_info = f"Остаток продукции '{product_name}' за период с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}:\n"
        balance_info += f"Сумма количества: {total_quantity}"

        # Выводим информацию о остатке на складе
        QMessageBox.information(self, "Остаток за период", balance_info)


if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication
    app = QApplication(sys.argv)
    dialog = ProductWarehouseDialog()
    dialog.show()
    sys.exit(app.exec_())
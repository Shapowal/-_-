import os
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QPushButton, QMessageBox, QInputDialog
from openpyxl import Workbook, load_workbook

class ProductCompositionDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Настройка состава продукта")
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Добавление кнопки "Добавить материал в продукт"
        add_material_button = QPushButton("Добавить материал в продукт")
        add_material_button.clicked.connect(self.add_material_to_product)
        layout.addWidget(add_material_button)

        # Проверяем и добавляем заголовки в production.xlsx
        self.ensure_production_file_has_headers()

        # Проверяем и добавляем заголовки в product_composition.xlsx
        self.ensure_composition_file_has_headers()

    def ensure_production_file_has_headers(self):
        filename = "production.xlsx"
        if not os.path.exists(filename):
            wb = Workbook()
            for sheet_name in ["Линия 1,5", "Линия 0,5", "Линия 1,5 сладкая", "Линия 0,5 сладкая", "Линия 5 литров", "Линия 19 литров"]:
                wb.create_sheet(sheet_name)
                ws = wb[sheet_name]
                ws.append(["Название продукта", "GTIN", "Скважина", "Тип продукта"])
            wb.save(filename)
            wb.close()
        else:
            wb = load_workbook(filename)
            for sheet_name in ["Линия 1,5", "Линия 0,5", "Линия 1,5 сладкая", "Линия 0,5 сладкая", "Линия 5 литров", "Линия 19 литров"]:
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                    ws.append(["Название продукта", "GTIN", "Скважина", "Тип продукта"])
            wb.save(filename)
            wb.close()

    def ensure_composition_file_has_headers(self):
        filename = "product_composition.xlsx"
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["Материал", "Количество"])  # Добавляем заголовки
            wb.save(filename)
            wb.close()
        else:
            wb = load_workbook(filename)
            ws = wb.active
            if "Материал" not in [cell.value for cell in ws[1]] or "Количество" not in [cell.value for cell in ws[1]]:
                ws.insert_rows(1)  # Вставляем новую строку для заголовков
                ws["A1"] = "Материал"
                ws["B1"] = "Количество"
                wb.save(filename)
            wb.close()

    def add_material_to_product(self):
        products = self.get_products()
        if not products:
            QMessageBox.warning(self, "Ошибка", "Список продуктов пуст.")
            return

        selected_product, ok = QInputDialog.getItem(self, "Выберите продукт", "Выберите продукт:", products, 0, False)
        if not ok:
            return

        materials = self.get_materials()
        if not materials:
            QMessageBox.warning(self, "Ошибка", "Список материалов пуст.")
            return

        selected_material, ok = QInputDialog.getItem(self, "Выберите материал", "Выберите материал:", materials, 0, False)
        if not ok:
            return

        quantity, ok = QInputDialog.getInt(self, "Введите количество", "Введите количество:", min=1)
        if not ok:
            return

        # Делаем что-то с выбранным продуктом, материалом и количеством
        QMessageBox.information(self, "Успех",
                                f"Добавлен материал '{selected_material}' в продукт '{selected_product}' в количестве {quantity}.")

    def get_products(self):
        # Здесь должна быть логика получения списка продуктов из файла или базы данных
        # Например, из файла production.xlsx
        filename = "production.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл с продуктами не найден.")
            return []

        wb = load_workbook(filename)
        products = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                products.extend(row)
        wb.close()
        return products

    def get_materials(self):
        # Здесь должна быть логика получения списка материалов со склада материалов
        # Например, из файла materials.xlsx
        filename = "materials.xlsx"
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл со складом материалов не найден.")
            return []

        wb = load_workbook(filename)
        materials = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                materials.extend(row)
        wb.close()
        return materials